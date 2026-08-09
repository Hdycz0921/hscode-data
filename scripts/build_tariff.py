# -*- coding: utf-8 -*-
"""
GitHub Actions 流水线 Step 1: 从 Excel 生成 p1~p4.json 税则分片数据

输入:
  sys.argv[1]  Excel 文件路径（可以是 data/source/xxx.xlsx 或绝对路径）

输出:
  scripts/tariff_output/p1.json  (第1-9章)
  scripts/tariff_output/p2.json  (第10-25章)
  scripts/tariff_output/p3.json  (第26-50章)
  scripts/tariff_output/p4.json  (第51-97章)

GitHub Actions 环境:
  actions/setup-python@v5  → Python 3.11
  pip install openpyxl     → 依赖已装

本地调试:
  python build_tariff.py "E:/报关用文件/2026年商品编码税率与申报要素表.xlsx"
"""
import sys, os, json, re

sys.stdout.reconfigure(encoding='utf-8')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(SCRIPT_DIR, 'tariff_output')
os.makedirs(OUT_DIR, exist_ok=True)


def get_shard(code10):
    """根据10位HS编码返回分片编号 1-4"""
    try:
        c = str(code10).zfill(10)
        ch = int(c[0:2])
    except:
        return 4
    if ch <= 9:
        return 1
    elif ch <= 25:
        return 2
    elif ch <= 50:
        return 3
    else:
        return 4


def sheet_row_iter(ws):
    """yield (row_idx, row_data) 跳过表头行"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    # 找表头行：第1列含 "编码" 或 "HS" 即为表头
    header_idx = 0
    for i, row in enumerate(rows):
        cell0 = str(row[0] if row[0] else '').strip()
        if re.search(r'编码|HS.*Code|code', cell0, re.IGNORECASE):
            header_idx = i
            break
    for i in range(header_idx + 1, len(rows)):
        row = rows[i]
        if not row or not row[0]:
            continue
        yield i, row


def parse_excel(excel_path):
    print(f'读取 Excel: {excel_path}')
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = None
    # 找主数据工作表
    for name in wb.sheetnames:
        if re.search(r'2026.*HS|HS.*编码.*申报|申报要素', name):
            ws = wb[name]
            print(f'  使用工作表: {name}')
            break
    if not ws:
        ws = wb.active
        print(f'  使用活动工作表: {ws.title}')

    shards = {1: [], 2: [], 3: [], 4: []}
    seen = set()
    skipped_empty = 0

    for row_idx, row in sheet_row_iter(ws):
        try:
            raw_code = str(row[0]).strip() if row[0] else ''
        except:
            continue

        # 跳过空行
        if not raw_code or raw_code in ('None', 'nan'):
            skipped_empty += 1
            continue

        # 标准化为10位HS编码
        code10 = raw_code.replace('.', '').replace(' ', '').zfill(10)
        code10 = re.sub(r'[^0-9]', '', code10).zfill(10)

        if len(code10) != 10 or not code10.isdigit():
            continue
        if code10 in seen:
            continue
        seen.add(code10)

        try:
            name = str(row[1]).strip() if row[1] else ''
        except:
            name = ''
        try:
            decl = str(row[3]).strip() if row[3] else ''
        except:
            decl = ''
        if decl in ('nan', 'None', ''):
            decl = ''

        # t1~t8: 列 E-L (索引 4-11)
        rates = []
        for col_idx in range(4, 12):
            try:
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    rates.append('')
                elif isinstance(val, (int, float)):
                    rates.append(str(val))
                else:
                    rates.append(str(val).strip())
            except:
                rates.append('')

        # 监管代码 N(列13), 检验检疫 O(列14), 法定单位 P(列15), 第二单位 Q(列16)
        try:
            sup = str(row[13]).strip() if row[13] else ''
        except:
            sup = ''
        try:
            insp = str(row[14]).strip() if row[14] else ''
        except:
            insp = ''
        try:
            unit = str(row[15]).strip() if row[15] else ''
        except:
            unit = ''
        try:
            unit2 = str(row[16]).strip() if row[16] else ''
        except:
            unit2 = ''

        if sup in ('nan', 'None'):
            sup = ''
        if insp in ('nan', 'None'):
            insp = ''
        if unit in ('nan', 'None'):
            unit = ''
        if unit2 in ('nan', 'None'):
            unit2 = ''

        item = {
            'code': code10,
            'name': name,
            'd': decl,
            't1': rates[0], 't2': rates[1], 't3': rates[2], 't4': rates[3],
            't5': rates[4], 't6': rates[5], 't7': rates[6], 't8': rates[7],
            's': sup,
            'i': insp,
            'u': unit,
            'u2': unit2,
        }
        shard = get_shard(code10)
        shards[shard].append(item)

    wb.close()
    return shards


def main():
    if len(sys.argv) < 2:
        print('用法: python build_tariff.py <Excel文件路径>')
        sys.exit(1)

    excel_path = sys.argv[1]

    # 支持相对路径（相对于仓库根目录）
    if not os.path.isabs(excel_path):
        repo_root = os.path.dirname(SCRIPT_DIR)
        excel_path = os.path.join(repo_root, excel_path)

    excel_path = os.path.abspath(excel_path)
    if not os.path.exists(excel_path):
        print(f'错误: 文件不存在 {excel_path}')
        sys.exit(1)

    shards = parse_excel(excel_path)
    total = sum(len(v) for v in shards.values())
    print(f'\n生成税则分片 (共 {total} 条):')

    for i in range(1, 5):
        out_path = os.path.join(OUT_DIR, f'p{i}.json')
        data = shards[i]
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        sz = os.path.getsize(out_path)
        print(f'  p{i}.json  {len(data):5d} 条  {sz/1024:.1f} KB  → {out_path}')

    print(f'\n✅ 税则数据生成完成: {OUT_DIR}')


if __name__ == '__main__':
    main()
